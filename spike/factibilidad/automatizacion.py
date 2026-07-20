#!/usr/bin/env python3
"""
Dashboard de popularidad de productos.

Toma un reporte de popularidad exportado del sistema del restaurante (una hoja
de Excel "sucia") y produce los datos de un dashboard: métricas clave, desglose
por familia y el top de productos.

El archivo real trae basura de sistema:
  - dos tablas apiladas en la misma hoja (el detalle y "Totales por familia"),
  - una fila "Total" con las columnas corridas,
  - encabezados de reporte y filas de resumen.

El script detecta la tabla de detalle por su encabezado y corta al llegar a la
segunda tabla; no depende de números de fila fijos, así funciona con cualquier
archivo del mismo formato.

Uso:
    python automatizacion.py <entrada.xlsx> [carpeta_salida]
"""

import json
import sys
from pathlib import Path

import openpyxl


# --- Localización de columnas por nombre de encabezado (tolerante) ----------

def _norm(x):
    return str(x).strip().lower() if x is not None else ""


def _encontrar_encabezado(filas):
    """Devuelve (indice_fila, mapa_columnas) de la tabla de DETALLE.

    La reconoce porque su encabezado tiene, a la vez, 'producto', 'ingreso' y
    'familia' — la fila de arranque de la segunda tabla no las tiene todas.
    """
    for i, fila in enumerate(filas):
        etiquetas = [_norm(c) for c in fila]
        tiene = lambda t: any(t == e or t in e for e in etiquetas)
        if tiene("producto") and tiene("ingreso") and tiene("familia"):
            mapa = {}
            for j, e in enumerate(etiquetas):
                if e in ("producto",) and "producto" not in mapa:
                    mapa["producto"] = j
                elif e == "familia" and "familia" not in mapa:
                    mapa["familia"] = j
                elif e == "cantidad" and "cantidad" not in mapa:
                    mapa["cantidad"] = j
                elif e == "ingreso" and "ingreso" not in mapa:
                    mapa["ingreso"] = j
                elif e == "costo" and "costo" not in mapa:
                    mapa["costo"] = j
                elif e == "utilidad" and "utilidad" not in mapa:
                    mapa["utilidad"] = j
            return i, mapa
    raise ValueError("No se encontró la tabla de detalle (falta encabezado con Producto/Ingreso/Familia).")


def _num(x):
    """Convierte a float tolerando None, texto con símbolos, o valores vacíos."""
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).replace("$", "").replace(",", "").replace("%", "").strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _es_marcador_resumen(fila):
    """True si la fila arranca otra tabla o es un total/subtotal de sistema."""
    textos = [_norm(c) for c in fila if isinstance(c, str)]
    for t in textos:
        if "totales por familia" in t or t == "total" or t.startswith("total "):
            return True
    return False


# --- Extracción del detalle -------------------------------------------------

def extraer_detalle(ruta):
    wb = openpyxl.load_workbook(ruta, data_only=True, read_only=True)
    ws = wb.active
    filas = [list(f) for f in ws.iter_rows(values_only=True)]

    inicio, col = _encontrar_encabezado(filas)

    productos = []
    for fila in filas[inicio + 1:]:
        # Fin del detalle: llegó la segunda tabla o un total de sistema.
        if _es_marcador_resumen(fila):
            break

        producto = fila[col["producto"]] if col["producto"] < len(fila) else None
        # Solo cuentan filas cuyo Producto es TEXTO real (las filas basura traen
        # número o vacío en esa columna por las columnas corridas).
        if not isinstance(producto, str) or not producto.strip():
            continue

        ingreso = _num(fila[col["ingreso"]]) if col.get("ingreso") is not None else None
        utilidad = _num(fila[col["utilidad"]]) if col.get("utilidad") is not None else None
        if ingreso is None or utilidad is None:
            continue  # sin cifras válidas → no es una fila de producto

        productos.append({
            "producto": producto.strip(),
            "familia": (str(fila[col["familia"]]).strip()
                        if col.get("familia") is not None and fila[col["familia"]] else "SIN FAMILIA"),
            "cantidad": _num(fila[col["cantidad"]]) if col.get("cantidad") is not None else 0.0,
            "ingreso": round(ingreso, 2),
            "costo": round(_num(fila[col["costo"]]) or (ingreso - utilidad), 2) if col.get("costo") is not None else round(ingreso - utilidad, 2),
            "utilidad": round(utilidad, 2),
        })

    return productos


# --- Armado del dashboard ---------------------------------------------------

def construir_dashboard(productos):
    ingreso_total = round(sum(p["ingreso"] for p in productos), 2)
    costo_total = round(sum(p["costo"] for p in productos), 2)
    utilidad_total = round(sum(p["utilidad"] for p in productos), 2)
    cantidad_total = round(sum(p["cantidad"] for p in productos), 2)

    por_familia = {}
    for p in productos:
        f = por_familia.setdefault(p["familia"], {"familia": p["familia"], "cantidad": 0.0, "ingreso": 0.0, "utilidad": 0.0})
        f["cantidad"] += p["cantidad"]
        f["ingreso"] += p["ingreso"]
        f["utilidad"] += p["utilidad"]
    familias = []
    for f in por_familia.values():
        f["ingreso"] = round(f["ingreso"], 2)
        f["utilidad"] = round(f["utilidad"], 2)
        f["cantidad"] = round(f["cantidad"], 2)
        f["margen_pct"] = round(100 * f["utilidad"] / f["ingreso"], 1) if f["ingreso"] else 0.0
        familias.append(f)
    familias.sort(key=lambda x: x["ingreso"], reverse=True)

    top = sorted(productos, key=lambda p: p["ingreso"], reverse=True)[:10]

    return {
        "metricas": {
            "num_productos": len(productos),
            "num_familias": len(familias),
            "cantidad_total": cantidad_total,
            "ingreso_total": ingreso_total,
            "costo_total": costo_total,
            "utilidad_total": utilidad_total,
            "margen_pct": round(100 * utilidad_total / ingreso_total, 1) if ingreso_total else 0.0,
        },
        "familias": familias,
        "top_productos": [
            {"producto": p["producto"], "familia": p["familia"], "cantidad": p["cantidad"],
             "ingreso": p["ingreso"], "utilidad": p["utilidad"]}
            for p in top
        ],
    }


def escribir_xlsx(dash, ruta):
    from openpyxl import Workbook
    wb = Workbook()
    m = wb.active
    m.title = "Metricas"
    for k, v in dash["metricas"].items():
        m.append([k, v])

    hf = wb.create_sheet("Familias")
    hf.append(["Familia", "Cantidad", "Ingreso", "Utilidad", "Margen %"])
    for f in dash["familias"]:
        hf.append([f["familia"], f["cantidad"], f["ingreso"], f["utilidad"], f["margen_pct"]])

    ht = wb.create_sheet("Top productos")
    ht.append(["Producto", "Familia", "Cantidad", "Ingreso", "Utilidad"])
    for p in dash["top_productos"]:
        ht.append([p["producto"], p["familia"], p["cantidad"], p["ingreso"], p["utilidad"]])

    wb.save(ruta)


MANIFIESTO = {
    "entradas": [{
        "nombre": "reporte",
        "tipo": "archivo",
        "formato": "xlsx",
        "descripcion": "Reporte de popularidad de productos exportado del sistema del "
                       "restaurante (una fila por producto: cantidad, ingreso, costo, "
                       "utilidad y familia).",
        "requerido": True,
    }]
}


def main():
    if len(sys.argv) < 2:
        print("Uso: python automatizacion.py <entrada.xlsx> [carpeta_salida]", file=sys.stderr)
        sys.exit(1)

    entrada = Path(sys.argv[1])
    salida = Path(sys.argv[2]) if len(sys.argv) > 2 else entrada.parent
    salida.mkdir(parents=True, exist_ok=True)

    productos = extraer_detalle(entrada)
    dash = construir_dashboard(productos)

    (salida / "resultado.json").write_text(
        json.dumps(dash, ensure_ascii=False, indent=2), encoding="utf-8")
    escribir_xlsx(dash, salida / "dashboard.xlsx")
    (salida / "manifiesto.json").write_text(
        json.dumps(MANIFIESTO, ensure_ascii=False, indent=2), encoding="utf-8")

    met = dash["metricas"]
    print(f"Productos:  {met['num_productos']}")
    print(f"Familias:   {met['num_familias']}")
    print(f"Ingreso:    {met['ingreso_total']:,.2f}")
    print(f"Utilidad:   {met['utilidad_total']:,.2f}")
    print(f"Margen:     {met['margen_pct']}%")
    print(f"Top 2:      {dash['top_productos'][0]['producto']} | {dash['top_productos'][1]['producto']}")
    print(f"Salidas en: {salida}")


if __name__ == "__main__":
    main()
